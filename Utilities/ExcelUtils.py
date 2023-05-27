from openpyxl import load_workbook


def getRowCount(file):
    workbook = load_workbook(file)
    sheet = workbook.active
    max_row = sheet.max_row
    return max_row


def getColumnCount(file):
    workbook = load_workbook(file)
    sheet = workbook.active
    max_column = sheet.max_column
    return max_column


def readData(file,rownum,columnnum):
    workbook = load_workbook(file)
    sheet = workbook.active
    return sheet.cell(row=rownum, column=columnnum).value


def writeData(file, rownum, columnnum, data):
    workbook = load_workbook(file)
    sheet = workbook.active
    sheet.cell(row=rownum, column=columnnum).value = data
    workbook.save(file)